from bs4 import BeautifulSoup
import requests
import re
import openpyxl

def makeSpreadsheet(stateRecords):
	book = openpyxl.Workbook()
	book.create_sheet('Sheet2')
	sheet = book.get_sheet_by_name('Sheet')
	sheet2 = book.get_sheet_by_name('Sheet2')
	row = 1
	column = 2
	total = ['0-0-0' for x in range(len(stateRecords))]
	for state in stateList:
		sheet.cell(row=row, column=column).value = state
		sheet2.cell(row=row, column=column).value = state
		column = column + 1
	row = row + 1
	for firstState in stateRecords:
		column = 1
		sheet.cell(row=row, column=column).value = stateList[row-2]
		sheet2.cell(row=row, column=column).value = stateList[row-2]
		column = column + 1
		wins = 0
		losses = 0
		ties = 0
		for secondState in firstState:	
			sheet.cell(row=row, column=column).value = secondState
			if secondState != '':
				getNumbers = secondState.split('-')
				sheet2.cell(row=row, column=column).value = round((int(getNumbers[0]) + int(getNumbers[2])/2)/(int(getNumbers[0])+int(getNumbers[1])+int(getNumbers[2])),3)
				wins = wins + int(getNumbers[0])
				losses = losses + int(getNumbers[1])
				ties = ties + int(getNumbers[2])
				currentRecord = total[column - 2].split('-')
				total[column - 2] = str(int(currentRecord[0]) + int(getNumbers[0])) + '-' + str(int(currentRecord[1]) + int(getNumbers[1])) + '-' + str(int(currentRecord[2]) + int(getNumbers[2]))
			else:
				sheet2.cell(row=row, column=column).value = secondState
			column = column + 1
		sheet.cell(row=row, column=column).value = str(wins) + '-' + str(losses) + '-' + str(ties)
		sheet2.cell(row=row, column=column).value = round((wins + ties/2)/(wins+losses+ties),3)
		row = row + 1
	column = 2
	for eachState in total:
		splitTotal = eachState.split('-')
		sheet.cell(row=row, column=column).value = eachState
		sheet2.cell(row=row, column=column).value = round((int(splitTotal[0])+int(splitTotal[2])/2)/(int(splitTotal[0])+int(splitTotal[1])+int(splitTotal[2])),3)
		column = column + 1
	book.save('stateRecords.xlsx')

def getState(teamName):
	teamFile = open('teamsStates.txt')
	for line in teamFile:
		splitLine = line.split(",")
		if splitLine[0] == teamName:
			return splitLine[1][:2]

def getTeamInfo(teamlink):
	#Get team state
	teamState = teamlink.split("/")
	teamState = teamState[3].replace('-', ' ').upper()
	firstState = getState(teamState)
	page = requests.get(teamlink)
	content = page.content
	soup = (BeautifulSoup(content, "html.parser"))
	teamSection = soup.find("div", {"class": "row-fluid matchupslist row4"})
	eachTeam = teamSection.find_all("li")
	for team in eachTeam:
		teamDiv = team.div
		if(teamDiv['data-conference'] != "NON-FBS TEAMS" and teamDiv['data-name'] != "Idaho"):
			#GET STATE OF TEAM
			secondState = getState(teamDiv['data-name'].replace('-', ' ').upper())
			if firstState != secondState:
				if stateRecords[stateList.index(firstState)][stateList.index(secondState)] == '':
					stateRecords[stateList.index(firstState)][stateList.index(secondState)] = str(teamDiv['data-winscount']) + '-' + str(teamDiv['data-lossescount']) + '-' + str(teamDiv['data-tiescount'])
				else:
					current = stateRecords[stateList.index(firstState)][stateList.index(secondState)].split('-')
					stateRecords[stateList.index(firstState)][stateList.index(secondState)] = str(int(teamDiv['data-winscount']) + int(current[0])) + '-' + str(int(teamDiv['data-lossescount']) + int(current[1])) + '-' + str(int(teamDiv['data-tiescount']) + int(current[2]))
			
if __name__ == "__main__":
	stateList = ['AL','AR','AZ','CA','CO','CT','FL','GA','HI','IA','ID','IL','IN','KS','KY','LA','MA','MD','MI','MN','MO','MS','NC','NE','NJ','NM','NV','NY','OH','OK','OR','PA','SC','TN','TX','UT','VA','WA','WI','WV','WY']
	stateRecords = [['' for x in range(len(stateList))] for y in range(len(stateList))]
	soup = BeautifulSoup(open("teampage.html"), "html.parser")
	teams = soup.find_all("li", "span4") + soup.find_all("li", "span3")
	teamURLs = []
	for li in teams:
		li = str(li)
		teampage = re.search('href="(.*)">', li)
		teamURLs.append(teampage.group(1))
	#getTeamInfo(teamURLs[0])
	for eachTeam in teamURLs:
		print(eachTeam)
		getTeamInfo(eachTeam)
	makeSpreadsheet(stateRecords)